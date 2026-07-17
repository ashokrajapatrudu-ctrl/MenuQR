#!/usr/bin/env python3
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path("/Users/ashokrajapatrudulake/Downloads")
SHEET = "05 Item-Level Master"

SOURCES = [
    ("Amigos", "AMIGOS_Designer_Ready_Menu_Content(1).xlsx", "AMIGOS_Designer_Ready_Menu_Content.xlsx"),
    ("Spicy Darbar", "SPICY_DARBAR_Designer_Ready_Menu_Content(1).xlsx", "SPICY_DARBAR_Designer_Ready_Menu_Content.xlsx"),
    ("Red Panda", "RED_PANDA_Designer_Ready_Menu_Content(1).xlsx", "RED_PANDA_Designer_Ready_Menu_Content.xlsx"),
    ("Amigos Grill Cafe", "AMIGOS_GRILL_CAFE_Designer_Ready_Menu_Content(1).xlsx", "AMIGOS_GRILL_CAFE_Designer_Ready_Menu_Content.xlsx"),
    ("Amigos Andhra Bhawan", "AMIGOS_ANDHRA_BHAWAN_Designer_Ready_Menu_Content(1).xlsx", "AMIGOS_ANDHRA_BHAWAN_Designer_Ready_Menu_Content.xlsx"),
]

REQUIRED_COLUMNS = [
    "Menu Section",
    "Product Name",
    "Portion",
    "Short Description",
    "Dietary",
    "Spice",
    "Main Ingredient Quantity (Approx.)",
    "Approx. Total Serving",
    "Serves",
    "Est. Calories",
    "Calorie Range",
    "Dine-In Price",
    "Recommended Badge(s)",
    "Likely Allergens*",
    "Validation / Owner Check",
    "Original Category",
    "Original Product Name",
]

ALLOWED_BADGES = {
    "signature": "Signature",
    "chef s special": "Chef's Special",
    "chef special": "Chef's Special",
    "most ordered": "Popular Pick",
    "popular pick": "Popular Pick",
    "best value": "Best Value",
    "veg favourite": "Veg Favourite",
    "veg favorite": "Veg Favourite",
    "spicy pick": "Spicy Pick",
}

BRAND_BY_CATEGORY = {
    "veg-starters": "Red Panda",
    "chicken-starters": "Red Panda",
    "seafood-starters": "Red Panda",
    "chinese-main-course": "Red Panda",
    "chinese-rice": "Red Panda",
    "noodles": "Red Panda",
    "soups": "Red Panda",
    "shawarma-wraps": "Amigos Grill Cafe",
    "mandi": "Amigos Grill Cafe",
    "burgers": "Amigos Grill Cafe",
    "continental": "Amigos Grill Cafe",
    "salads": "Amigos Grill Cafe",
    "thick-shakes": "Amigos Grill Cafe",
    "mocktails": "Amigos Grill Cafe",
    "south-indian": "Amigos Andhra Bhawan",
    "nonveg-dosa": "Amigos Andhra Bhawan",
    "veg-main-course": "Spicy Darbar",
    "paneer-main-course": "Spicy Darbar",
    "egg-main-course": "Spicy Darbar",
    "chicken-main-course": "Spicy Darbar",
    "mutton-main-course": "Spicy Darbar",
    "seafood-main-course": "Spicy Darbar",
    "tandoori-kebabs": "Spicy Darbar",
    "rotis-naans": "Spicy Darbar",
    "dals": "Spicy Darbar",
    "cold-beverages": "Spicy Darbar",
    "veg-biryanis": "Amigos",
    "veg-pot-biryanis": "Amigos",
    "nonveg-biryanis": "Amigos",
    "nonveg-pot-biryanis": "Amigos",
    "family-packs": "Amigos",
    "biryani-addons": "Amigos",
}


def resolve_source(primary, fallback):
    candidates = [
        ROOT / "data" / "menu-source" / primary,
        ROOT / "data" / "menu-source" / fallback,
        DOWNLOADS / primary,
        DOWNLOADS / fallback,
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError(f"Missing approved workbook: {primary} or {fallback}")


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize(value):
    text = clean_text(value).lower()
    text = text.replace("&", " and ")
    text = text.replace("tangadi", "tangdi")
    text = text.replace("boondhi", "boondi")
    text = re.sub(r"\bwith bone\b", " bone in ", text)
    text = re.sub(r"\bbn\b", " bone in ", text)
    text = re.sub(r"\bbl\b", " boneless ", text)
    text = re.sub(r"\bhalf\b", " half portion ", text)
    text = re.sub(r"\bfull\b", " full portion ", text)
    text = re.sub(r"\bsingle unit price\b|\bsingle serving\b", " single ", text)
    text = re.sub(r"\bregular\b", " regular ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def product_name_key(value):
    text = normalize(value)
    text = re.sub(r"\bselected days only\b|\bselected days\b|\bcalendar selected days\b", " ", text)
    text = re.sub(r"\b2 pieces\b|\b2 piece\b|\b2 pcs\b|\b4 pieces\b|\b4 piece\b|\b4 pcs\b|\b8 pieces\b|\b8 piece\b|\b8 pcs\b", " ", text)
    text = re.sub(r"\bloaded cheese mac and cheese\b", "loaded mac and cheese", text)
    text = re.sub(r"\bjuicy chicken mandi\b", "chicken juicy mandi", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def compact(value):
    return product_name_key(value).replace(" ", "")


def portion_key(value):
    text = normalize(value)
    if "online portion half" in text or "online half" in text:
        return "online half"
    if "8 pcs" in text or "8 pieces" in text or "8 piece" in text:
        return "8 pcs"
    if "4 pcs" in text or "4 pieces" in text or "4 piece" in text:
        return "4 pcs"
    if "2 pcs" in text or "2 pieces" in text or "2 piece" in text:
        return "2 pcs"
    if "half chicken platter" in text:
        return "half"
    if "whole chicken platter" in text:
        return "full"
    if "family platter" in text or "family pack" in text:
        return "family"
    if "full platter" in text:
        return "full"
    if "half platter" in text:
        return "half"
    if "quarter" in text:
        return "quarter"
    if "pot" in text:
        return "pot"
    if "add on" in text or "addon" in text:
        return "regular"
    if "half portion" in text:
        return "half"
    if "full portion" in text:
        return "full"
    if "300 ml" in text:
        return "300 ml"
    if "single" in text:
        return "single"
    if "regular" in text:
        return "regular"
    return text


def website_brand(item):
    name = normalize(item["name"])
    category = item["category"]
    if category == "meals-thalis":
        return "Spicy Darbar" if "north indian" in name else "Amigos Andhra Bhawan"
    if category == "indian-rice":
        return "Amigos Andhra Bhawan" if any(term in name for term in ["curd rice", "lemon rice", "tomato rice"]) else "Spicy Darbar"
    return BRAND_BY_CATEGORY.get(category, "Amigos")


def price_value(value):
    if value is None or value == "":
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def parse_diet(value):
    text = normalize(value)
    if "seafood" in text:
        return {"key": "seafood", "label": "Seafood"}
    if "egg" in text:
        return {"key": "egg", "label": "Egg"}
    if "non veg" in text or "nonveg" in text:
        return {"key": "nonveg", "label": "Non-Veg"}
    return {"key": "veg", "label": "Veg"}


def parse_spice(value):
    text = normalize(value)
    if "hot" in text:
        return "Hot"
    if "medium" in text:
        return "Medium"
    if "mild" in text:
        return "Mild"
    return ""


def parse_badges(value):
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"[|,;/]+", text)
    badges = []
    for part in parts:
        normalized = normalize(part)
        normalized = re.sub(r"^[a-z] ", "", normalized).strip()
        label = ALLOWED_BADGES.get(normalized)
        if label and label not in badges:
            badges.append(label)
    return badges[:2]


def allergen_copy(value):
    text = clean_text(value)
    normalized = normalize(text)
    if not text:
        return ""
    if "recipe validation required" in normalized:
        return "Allergen information requires confirmation from our team."
    if "none specifically identified" in normalized:
        return "No major allergen specifically identified. Please confirm with our team if you have an allergy."
    allergens = []
    for raw in re.split(r"[,;/]+", text):
        item = normalize(raw)
        if not item:
            continue
        if item == "soy":
            allergens.append("Contains soy")
        elif item == "dairy":
            allergens.append("Contains dairy")
        elif item == "gluten":
            allergens.append("Contains gluten")
        elif item == "egg":
            allergens.append("Contains egg")
        elif item in {"nuts", "nut"}:
            allergens.append("Contains nuts")
        elif item == "shellfish":
            allergens.append("Contains shellfish")
        elif item == "fish":
            allergens.append("Contains fish")
        elif item == "sesame":
            allergens.append("Contains sesame")
        else:
            allergens.append(f"Contains {raw.strip().lower()}")
    return "; ".join(dict.fromkeys(allergens))


def stable_id(brand, product, portion, price=None):
    return "-".join(part for part in [
        normalize(brand).replace(" ", "-"),
        normalize(product).replace(" ", "-"),
        normalize(portion).replace(" ", "-"),
        f"rs-{price}" if price is not None else "",
    ] if part)


def read_menu_data():
    source = (ROOT / "menu-data.js").read_text()
    match = re.search(r"window\.MENU_DATA\s*=\s*(\{.*\});\s*$", source, re.S)
    if not match:
        raise RuntimeError("Could not parse menu-data.js")
    return json.loads(match.group(1))


def read_workbooks():
    records = []
    source_paths = {}
    for brand, primary, fallback in SOURCES:
        path = resolve_source(primary, fallback)
        source_paths[brand] = str(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[SHEET]
        header = [cell.value for cell in next(worksheet.iter_rows(min_row=4, max_row=4))]
        indexes = {heading: index for index, heading in enumerate(header) if heading}
        missing = [column for column in REQUIRED_COLUMNS if column not in indexes]
        if missing:
            raise RuntimeError(f"{path.name} is missing columns: {', '.join(missing)}")
        for row_number, row in enumerate(worksheet.iter_rows(min_row=5, values_only=True), start=5):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            record = {column: row[indexes[column]] if indexes[column] < len(row) else None for column in REQUIRED_COLUMNS}
            record.update({
                "Brand": brand,
                "Workbook": path.name,
                "Sheet Row": row_number,
                "Dine-In Price": price_value(record["Dine-In Price"]),
                "_product_norm": normalize(record["Product Name"]),
                "_product_key": product_name_key(record["Product Name"]),
                "_product_compact": compact(record["Product Name"]),
                "_original_norm": normalize(record["Original Product Name"]),
                "_original_key": product_name_key(record["Original Product Name"]),
                "_original_compact": compact(record["Original Product Name"]),
                "_portion_key": portion_key(record["Portion"]),
                "_section_norm": normalize(record["Menu Section"]),
                "_category_norm": normalize(record["Original Category"]),
            })
            records.append(record)
    return records, source_paths


def build_indexes(records):
    indexes = defaultdict(list)
    for record in records:
        brand = record["Brand"]
        for field in ["_original_norm", "_product_norm", "_original_key", "_product_key", "_original_compact", "_product_compact"]:
            indexes[(brand, field, record[field], record["_portion_key"])].append(record)
            indexes[(brand, field, record[field], "")].append(record)
    return indexes


def candidate_score(item, variant, record):
    item_name = product_name_key(item["name"])
    item_compact = compact(item["name"])
    variant_portion = portion_key(f"{variant.get('label', '')} {variant.get('portion', '')}")
    variant_price = price_value(variant.get("price"))
    record_names = [
        ("Original Product Name", record["_original_key"], record["_original_compact"]),
        ("Product Name", record["_product_key"], record["_product_compact"]),
    ]

    name_score = 0
    name_method = ""
    for label, normalized_name, compact_name in record_names:
        if item_name == normalized_name:
            name_score = max(name_score, 100)
            name_method = f"exact {label.lower()}"
        elif item_compact == compact_name:
            name_score = max(name_score, 98)
            name_method = f"normalised {label.lower()}"
        else:
            ratio = max(
                SequenceMatcher(None, item_name, normalized_name).ratio(),
                SequenceMatcher(None, item_compact, compact_name).ratio(),
            )
            score = int(round(ratio * 100))
            if score > name_score:
                name_score = score
                name_method = f"fuzzy {label.lower()}"

    portion_score = 100 if variant_portion == record["_portion_key"] else 0
    if variant_portion in {"single", "regular"} and record["_portion_key"] in {"single", "regular", "300 ml"}:
        portion_score = 94
    if item["category"] == "mandi" and variant_portion == record["_portion_key"]:
        portion_score = max(portion_score, 92)
    if len(item.get("variants", [])) == 1 and not record["_portion_key"]:
        portion_score = max(portion_score, 90)
    price_score = 8 if variant_price is not None and variant_price == record["Dine-In Price"] else 0
    if name_score >= 96 and price_score and portion_score == 0:
        portion_score = 88

    category = normalize(item.get("_category_name", ""))
    category_score = 4 if category and (category == record["_section_norm"] or category == record["_category_norm"]) else 0
    confidence = min(100, int(round((name_score * 0.74) + (portion_score * 0.18) + category_score + price_score)))
    if portion_score == 0 and not price_score:
        confidence = min(confidence, 88)
    method = name_method
    if portion_score >= 94:
        method += " + portion"
    elif portion_score:
        method += " + inferred portion"
    return confidence, method


def match_variant(item, variant, records_by_brand):
    brand = website_brand(item)
    candidates = records_by_brand[brand]
    best = None
    for record in candidates:
        confidence, method = candidate_score(item, variant, record)
        price_match = price_value(variant.get("price")) == record["Dine-In Price"]
        if best is None or confidence > best[0] or (confidence == best[0] and price_match and price_value(variant.get("price")) != best[2]["Dine-In Price"]):
            best = (confidence, method, record)
    if best and best[0] >= 90:
        return brand, best[2], best[1], best[0], []

    low = []
    if best:
        low.append({
            "brand": brand,
            "existingWebsiteDishName": item["name"],
            "portion": variant.get("label") or variant.get("portion"),
            "spreadsheetProductName": best[2]["Product Name"],
            "matchMethod": best[1],
            "matchConfidence": best[0],
        })
    return brand, None, "unmatched", best[0] if best else 0, low


def validation_fields(note):
    text = normalize(note)
    fields = []
    checks = [
        ("Portion", ["portion", "platter", "piece count"]),
        ("Ingredient quantity", ["ingredient", "quantity"]),
        ("Serving size", ["serving", "serve", "size"]),
        ("Recipe", ["recipe", "proprietary", "rename", "spelling"]),
        ("Allergens", ["allergen"]),
        ("Badge", ["badge"]),
        ("Availability", ["availability", "available", "selected days", "calendar"]),
    ]
    for label, needles in checks:
        if any(needle in text for needle in needles):
            fields.append(label)
    return fields or ["Owner check"]


def main():
    menu = read_menu_data()
    category_names = {category["id"]: category["name"] for category in menu["categories"]}
    for item in menu["items"]:
        item["_category_name"] = category_names.get(item["category"], "")

    records, source_paths = read_workbooks()
    records_by_brand = defaultdict(list)
    for record in records:
        records_by_brand[record["Brand"]].append(record)

    details = {}
    matched_record_keys = set()
    match_details = []
    low_confidence = []
    price_mismatches = []
    exact_matches = 0
    normalised_matches = 0
    descriptions_updated = 0
    matched_items = set()

    for item in menu["items"]:
        item_detail = None
        variant_details = []
        item_brand = website_brand(item)
        item_badges = []
        item_spice = ""
        item_dietary = None
        item_description = ""
        for variant in item["variants"]:
            brand, record, method, confidence, low = match_variant(item, variant, records_by_brand)
            low_confidence.extend(low)
            updated = record is not None
            if record:
                matched_record_keys.add((record["Brand"], record["Workbook"], record["Sheet Row"]))
                matched_items.add(item["id"])
                if method.startswith("exact"):
                    exact_matches += 1
                else:
                    normalised_matches += 1
                if price_value(variant.get("price")) != record["Dine-In Price"]:
                    price_mismatches.append({
                        "brand": brand,
                        "dish": item["name"],
                        "portion": variant.get("label"),
                        "websitePrice": price_value(variant.get("price")),
                        "approvedDineInPrice": record["Dine-In Price"],
                    })
                description = clean_text(record["Short Description"])
                if description and not item_description:
                    item_description = description
                    descriptions_updated += 1
                badges = parse_badges(record["Recommended Badge(s)"])
                for badge in badges:
                    if badge not in item_badges:
                        item_badges.append(badge)
                if not item_spice:
                    item_spice = parse_spice(record["Spice"])
                if item_dietary is None:
                    item_dietary = parse_diet(record["Dietary"])
                variant_details.append({
                    "stableId": stable_id(brand, record["Product Name"], record["Portion"], record["Dine-In Price"]),
                    "label": clean_text(variant.get("label")),
                    "approvedPortion": clean_text(record["Portion"]),
                    "price": record["Dine-In Price"],
                    "mainIngredientQuantity": clean_text(record["Main Ingredient Quantity (Approx.)"]),
                    "approxTotalServing": clean_text(record["Approx. Total Serving"]),
                    "serves": clean_text(record["Serves"]),
                    "estimatedCalories": clean_text(record["Est. Calories"]),
                    "calorieRange": clean_text(record["Calorie Range"]),
                    "allergens": allergen_copy(record["Likely Allergens*"]),
                })
                match_details.append({
                    "brand": brand,
                    "existingWebsiteDishName": item["name"],
                    "spreadsheetProductName": clean_text(record["Product Name"]),
                    "portion": clean_text(record["Portion"]),
                    "matchMethod": method,
                    "matchConfidence": confidence,
                    "updated": True,
                })
            else:
                match_details.append({
                    "brand": brand,
                    "existingWebsiteDishName": item["name"],
                    "spreadsheetProductName": None,
                    "portion": clean_text(variant.get("label") or variant.get("portion")),
                    "matchMethod": method,
                    "matchConfidence": confidence,
                    "updated": False,
                })
        if variant_details:
            details[item["id"]] = {
                "stableProductId": stable_id(item_brand, item["name"], ""),
                "brand": item_brand,
                "menuSection": item["_category_name"],
                "description": item_description,
                "dietary": item_dietary or parse_diet(item.get("diet")),
                "spice": item_spice,
                "badges": item_badges[:2],
                "searchText": " ".join(filter(None, [
                    item_description,
                    item["_category_name"],
                    item_dietary["label"] if item_dietary else "",
                    item_spice,
                    " ".join(v["mainIngredientQuantity"] for v in variant_details),
                ])),
                "variants": variant_details,
            }

    unmatched_items = [
        {"id": item["id"], "name": item["name"], "category": item["category"], "brand": website_brand(item)}
        for item in menu["items"]
        if item["id"] not in matched_items
    ]
    unused_records = [
        {
            "brand": record["Brand"],
            "workbook": record["Workbook"],
            "sheetRow": record["Sheet Row"],
            "productName": clean_text(record["Product Name"]),
            "portion": clean_text(record["Portion"]),
            "menuSection": clean_text(record["Menu Section"]),
        }
        for record in records
        if (record["Brand"], record["Workbook"], record["Sheet Row"]) not in matched_record_keys
    ]

    validation = []
    seen_validation = set()
    for record in records:
        note = clean_text(record["Validation / Owner Check"])
        allergen = normalize(record["Likely Allergens*"])
        if not note and "recipe validation required" not in allergen:
            continue
        key = (record["Brand"], record["Product Name"], record["Portion"], note, record["Likely Allergens*"])
        if key in seen_validation:
            continue
        seen_validation.add(key)
        validation.append({
            "brand": record["Brand"],
            "productName": clean_text(record["Product Name"]),
            "portion": clean_text(record["Portion"]),
            "menuSection": clean_text(record["Menu Section"]),
            "fieldsRequiringConfirmation": validation_fields(note or record["Likely Allergens*"]),
            "ownerCheck": note or "Allergen information requires confirmation from our team.",
        })

    image_review = []
    for item in menu["items"]:
        image = clean_text(item.get("image"))
        if not image or image.endswith("fallback.webp"):
            image_review.append({
                "id": item["id"],
                "name": item["name"],
                "image": image,
                "reason": "Missing or fallback image; review against approved product description.",
            })

    import_report = {
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "sources": source_paths,
        "totalSpreadsheetRecords": len(records),
        "recordsByBrand": dict(Counter(record["Brand"] for record in records)),
        "websiteDishCount": len(menu["items"]),
        "websiteVariantCount": sum(len(item["variants"]) for item in menu["items"]),
        "matchedWebsiteDishes": len(matched_items),
        "descriptionsUpdated": descriptions_updated,
        "portionGroupsCreated": sum(1 for item in menu["items"] if item["id"] in matched_items and len(item["variants"]) > 1),
        "exactMatches": exact_matches,
        "normalisedMatches": normalised_matches,
        "unmatchedWebsiteDishes": unmatched_items,
        "unusedSpreadsheetRecords": unused_records,
        "lowConfidenceMatches": low_confidence,
        "priceMismatches": price_mismatches,
        "matchDetails": match_details,
    }

    product_details = {
        "meta": {
            "generatedAt": import_report["generatedAt"],
            "sourceSheet": SHEET,
            "totalProducts": len(details),
            "totalPortionDetails": sum(len(detail["variants"]) for detail in details.values()),
            "disclaimer": "Portion quantities and calorie values are approximate and may vary by recipe, preparation and actual serving. Please inform our team about allergies before ordering.",
        },
        "items": details,
    }

    js = "window.MENU_PRODUCT_DETAILS = "
    js += json.dumps(product_details, ensure_ascii=False, indent=2)
    js += ";\n"
    (ROOT / "menu-product-details.js").write_text(js)
    (ROOT / "menu-import-report.json").write_text(json.dumps(import_report, ensure_ascii=False, indent=2) + "\n")
    (ROOT / "menu-owner-validation.json").write_text(json.dumps({
        "generatedAt": import_report["generatedAt"],
        "totalItemsRequiringValidation": len(validation),
        "items": validation,
    }, ensure_ascii=False, indent=2) + "\n")
    (ROOT / "image-review-required.json").write_text(json.dumps({
        "generatedAt": import_report["generatedAt"],
        "totalItemsRequiringImageReview": len(image_review),
        "items": image_review,
        "note": "Existing approved image mapping was preserved. This report only flags missing or fallback images.",
    }, ensure_ascii=False, indent=2) + "\n")

    audit = [
        "# Menu Data Audit",
        "",
        f"Generated: {import_report['generatedAt']}",
        "",
        "## Source Workbooks",
        "",
    ]
    for brand, path in source_paths.items():
        audit.append(f"- {brand}: `{Path(path).name}`")
    audit.extend([
        "",
        "## Summary",
        "",
        f"- Total spreadsheet records processed: {len(records)}",
        f"- Website dishes matched: {len(matched_items)} of {len(menu['items'])}",
        f"- Website variants matched: {exact_matches + normalised_matches} of {import_report['websiteVariantCount']}",
        f"- Exact variant matches: {exact_matches}",
        f"- Normalised/fuzzy accepted variant matches: {normalised_matches}",
        f"- Descriptions updated from approved records: {descriptions_updated}",
        f"- Portion groups with approved variant details: {import_report['portionGroupsCreated']}",
        f"- Unmatched website dishes: {len(unmatched_items)}",
        f"- Low-confidence rejected matches: {len(low_confidence)}",
        f"- Owner validation items: {len(validation)}",
        f"- Price mismatches: {len(price_mismatches)}",
        "",
        "## Notes",
        "",
        "- The current website product list and prices were not changed.",
        "- The five brands were kept separate during matching; brand was inferred from existing website category ownership where the site did not already store a brand field.",
        "- Takeaway and Swiggy/Zomato columns were ignored and are not exported to the website.",
        "- Internal source rows, original category/product names and validation notes are not exported to customer-facing product details.",
        "- `AMIGOS_Designer_Ready_Menu_Content.xlsx` contains the full 527-row master, so many records from that workbook are intentionally unused when a focused brand workbook is the owner of that website category.",
    ])
    if unmatched_items:
        audit.extend(["", "## Unmatched Website Dishes", ""])
        audit.extend(f"- {item['brand']} / {item['name']} (`{item['id']}`)" for item in unmatched_items[:80])
    if price_mismatches:
        audit.extend(["", "## Price Mismatches", ""])
        audit.extend(f"- {item['brand']} / {item['dish']} / {item['portion']}: website ₹{item['websitePrice']} vs approved ₹{item['approvedDineInPrice']}" for item in price_mismatches[:80])
    (ROOT / "menu-data-audit.md").write_text("\n".join(audit) + "\n")

    print(json.dumps({
        "totalSpreadsheetRecords": len(records),
        "matchedWebsiteDishes": len(matched_items),
        "unmatchedWebsiteDishes": len(unmatched_items),
        "lowConfidenceMatches": len(low_confidence),
        "ownerValidationItems": len(validation),
        "descriptionsUpdated": descriptions_updated,
        "portionGroupsCreated": import_report["portionGroupsCreated"],
        "priceMismatches": len(price_mismatches),
    }, indent=2))


if __name__ == "__main__":
    main()
